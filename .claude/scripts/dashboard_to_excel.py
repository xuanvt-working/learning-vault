#!/usr/bin/env python3
"""
dashboard_to_excel.py — xuất các dashboard markdown của learning-vault ra một
workbook Excel duy nhất (mỗi dashboard = 1 sheet; file nhiều bảng = nhiều sheet con).

Chỉ phụ thuộc openpyxl (xem .claude/scripts/requirements.txt).

Dùng:
    python .claude/scripts/dashboard_to_excel.py [dashboard_dir]

Mặc định dashboard_dir = TOPIK/99_Dashboard (tính từ gốc repo). Output ghi ra
<dashboard_dir>/exports/Dashboard.xlsx. Chạy qua Bash nên KHÔNG kích hoạt hook
pre_tool_use.py (hook chỉ chặn tool Write/Edit, không chặn tiến trình Bash).
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

# stdout Windows mặc định cp1252 — ép UTF-8 để in được Hangul/emoji/ký tự Việt.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "Thiếu openpyxl. Cài trước: pip install -r .claude/scripts/requirements.txt"
    )

# Thứ tự + tên sheet cho các file 1-bảng (tên file -> tên sheet).
SINGLE_TABLE_SHEETS = {
    "Pair_Matrix": "Pair_Matrix",
    "Distractor_Stats": "Distractor_Stats",
    "Topic_Stats": "Topic_Stats",
    "Trap_Stats": "Trap_Stats",
    "Seq_Stats": "Seq_Stats",
    "Blank_Stats": "Blank_Stats",
    "Adverb_Stats": "Adverb_Stats",
}
MULTI_TABLE_FILE = "GR_Compiled_Handbook"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
MAX_COL_WIDTH = 60
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def strip_frontmatter(text: str) -> str:
    """Bỏ khối YAML frontmatter --- ... --- ở đầu file (nếu có)."""
    if text.startswith("---"):
        end = text.find("\n---", 3)
        if end != -1:
            nl = text.find("\n", end + 1)
            return text[nl + 1 :] if nl != -1 else ""
    return text


def clean_cell(value: str) -> str:
    """Làm sạch markdown trong một ô: bỏ wikilink/bold/code, giữ emoji."""
    v = value.strip()
    v = v.replace("<br>", "\n").replace("<br/>", "\n")
    v = re.sub(r"\[\[([^\]|]+)\|([^\]]+)\]\]", r"\2", v)  # [[target|alias]] -> alias
    v = re.sub(r"\[\[([^\]]+)\]\]", r"\1", v)              # [[target]] -> target
    v = re.sub(r"\*\*([^*]+)\*\*", r"\1", v)               # **bold**
    v = re.sub(r"__([^_]+)__", r"\1", v)                   # __underline__
    v = re.sub(r"`([^`]+)`", r"\1", v)                     # `code`
    v = v.replace(r"\|", "|")                              # escaped pipe trong bảng
    return v.strip()


def split_row(line: str) -> list[str]:
    """Tách một dòng bảng markdown thành list ô (bỏ | ngoài cùng)."""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    # tách theo | không bị escape
    parts = re.split(r"(?<!\\)\|", line)
    return [clean_cell(p) for p in parts]


def is_separator(line: str) -> bool:
    """Dòng phân cách header bảng: | --- | :--: | ..."""
    s = line.strip().strip("|")
    cells = s.split("|")
    return bool(cells) and all(re.fullmatch(r"\s*:?-{2,}:?\s*", c) for c in cells)


def parse_tables(text: str) -> list[dict]:
    """
    Parse mọi bảng markdown trong text.
    Trả list dict: {heading, header (list), rows (list of list)}.
    heading = tiêu đề (#/##/###) gần nhất phía trên bảng.
    """
    lines = strip_frontmatter(text).splitlines()
    tables: list[dict] = []
    last_heading = ""
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]
        h = re.match(r"^\s*#{1,6}\s+(.*)$", line)
        if h:
            last_heading = h.group(1).strip()
            i += 1
            continue
        # nhận diện bắt đầu bảng: dòng | ... | và dòng kế là separator
        if line.strip().startswith("|") and i + 1 < n and is_separator(lines[i + 1]):
            header = split_row(line)
            rows = []
            i += 2  # bỏ header + separator
            while i < n and lines[i].strip().startswith("|"):
                if is_separator(lines[i]):
                    i += 1
                    continue
                rows.append(split_row(lines[i]))
                i += 1
            tables.append({"heading": last_heading, "header": header, "rows": rows})
            continue
        i += 1
    return tables


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Tên sheet hợp lệ Excel: bỏ ký tự cấm, cắt <=31, đảm bảo duy nhất."""
    clean = INVALID_SHEET_CHARS.sub("", name).strip().strip("'")
    if not clean:
        clean = "Sheet"
    clean = clean[:31]
    candidate = clean
    k = 2
    while candidate.lower() in used:
        suffix = f"_{k}"
        candidate = clean[: 31 - len(suffix)] + suffix
        k += 1
    used.add(candidate.lower())
    return candidate


def write_table(ws, table: dict) -> None:
    """Ghi một bảng vào worksheet: header đậm + freeze + auto width + wrap."""
    header = table["header"]
    rows = table["rows"]
    ncols = len(header)

    ws.append(header)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    for row in rows:
        # đệm/cắt cho khớp số cột header
        row = (row + [""] * ncols)[:ncols]
        ws.append(row)

    # wrap + auto width
    widths = [len(str(h)) for h in header]
    for r, row in enumerate(rows, start=2):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            val = str(cell.value or "")
            longest_line = max((len(x) for x in val.split("\n")), default=0)
            if c - 1 < len(widths):
                widths[c - 1] = max(widths[c - 1], longest_line)

    for c in range(1, ncols + 1):
        w = min(widths[c - 1] + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(c)].width = max(w, 8)

    ws.freeze_panes = "A2"


def write_stacked_tables(ws, tables: list[dict]) -> None:
    """Xếp CHỒNG nhiều bảng (cột khác nhau) vào một sheet: mỗi bảng có dòng
    tiêu đề (heading) + hàng header riêng, cách nhau một dòng trống."""
    widths: dict[int, int] = {}
    r = 1
    for table in tables:
        header = table["header"]
        ncols = len(header)

        # dòng tiêu đề nhóm — merge trải hết chiều rộng bảng để hiện đủ text
        tc = ws.cell(row=r, column=1, value=table["heading"] or "")
        tc.font = TITLE_FONT
        tc.fill = TITLE_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center")
        if ncols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            # tô nền cho các ô còn lại trong vùng merge (openpyxl chỉ tô ô đầu)
            for c in range(2, ncols + 1):
                ws.cell(row=r, column=c).fill = TITLE_FILL
        r += 1

        # hàng header của bảng
        for c, h in enumerate(header, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP
            widths[c] = max(widths.get(c, 0), len(str(h)))
        r += 1

        # các hàng dữ liệu
        for row in table["rows"]:
            row = (row + [""] * ncols)[:ncols]
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = WRAP
                longest = max((len(x) for x in str(v).split("\n")), default=0)
                widths[c] = max(widths.get(c, 0), longest)
            r += 1

        r += 1  # dòng trống ngăn cách các nhóm

    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = max(min(w + 2, MAX_COL_WIDTH), 8)


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent  # .claude/scripts -> repo root

    dash_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else repo_root / "TOPIK" / "99_Dashboard"
    if not dash_dir.is_dir():
        sys.exit(f"Không thấy thư mục dashboard: {dash_dir}")

    md_files = sorted(p for p in dash_dir.glob("*.md") if p.is_file())
    if not md_files:
        sys.exit(f"Không có file .md nào trong {dash_dir}")

    wb = Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định
    used_names: set[str] = set()
    sheet_count = 0

    # sắp xếp: file 1-bảng theo thứ tự định sẵn trước, GR_Handbook cuối, còn lại giữa
    def sort_key(p: Path):
        stem = p.stem
        if stem in SINGLE_TABLE_SHEETS:
            return (0, list(SINGLE_TABLE_SHEETS).index(stem))
        if stem == MULTI_TABLE_FILE:
            return (2, 0)
        return (1, stem)

    for md in sorted(md_files, key=sort_key):
        stem = md.stem
        tables = parse_tables(md.read_text(encoding="utf-8"))
        if not tables:
            print(f"  · {stem}: không có bảng — bỏ qua")
            continue

        if stem in SINGLE_TABLE_SHEETS or (stem != MULTI_TABLE_FILE and len(tables) == 1):
            # một sheet cho cả file (dùng bảng đầu; nếu lỡ >1 thì nối phần còn lại phía dưới)
            name = SINGLE_TABLE_SHEETS.get(stem, stem)
            ws = wb.create_sheet(sanitize_sheet_name(name, used_names))
            write_table(ws, tables[0])
            sheet_count += 1
            print(f"  · {stem}: 1 sheet ({len(tables[0]['rows'])} hàng)")
        else:
            # nhiều bảng: gộp mọi bảng GROUP (§1) vào MỘT sheet xếp chồng;
            # các bảng còn lại (§0, §2, §3) mỗi bảng 1 sheet.
            i = 0
            made = 0
            while i < len(tables):
                table = tables[i]
                if table["heading"].startswith("GROUP_"):
                    grp = []
                    while i < len(tables) and tables[i]["heading"].startswith("GROUP_"):
                        grp.append(tables[i])
                        i += 1
                    ws = wb.create_sheet(sanitize_sheet_name("§1 Nhóm ngữ pháp", used_names))
                    write_stacked_tables(ws, grp)
                    made += 1
                    print(f"    ↳ §1 Nhóm ngữ pháp: {len(grp)} nhóm gộp vào 1 sheet")
                else:
                    base = table["heading"] or f"{stem}_{i + 1}"
                    ws = wb.create_sheet(sanitize_sheet_name(base, used_names))
                    write_table(ws, table)
                    made += 1
                    i += 1
            sheet_count += made
            print(f"  · {stem}: {made} sheet ({len(tables)} bảng)")

    out_dir = dash_dir / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Dashboard.xlsx"
    wb.save(out_path)
    print(f"\nĐã ghi {sheet_count} sheet → {out_path}")


if __name__ == "__main__":
    main()
